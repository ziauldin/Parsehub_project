"""
Excel Import Service for Metadata Management
Handles parsing and importing metadata from Excel files
"""

import os
import json
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pandas as pd
except ImportError:
    pd = None

from backend.database import ParseHubDatabase


class ExcelImportService:
    """Service for importing project metadata from Excel files"""
    
    # Expected columns in the Excel file
    EXPECTED_COLUMNS = {
        'Personal Project ID': 'personal_project_id',
        'Project ID (ParseHub)': 'project_token',
        'Project_name': 'project_name',
        'Last_run_data': 'last_run_date',
        'Create_date': 'created_date',
        'update_date': 'updated_date',
        'Region': 'region',
        'Country': 'country',
        'Brand': 'brand',
        'Website_url': 'website_url',
        'Total_pages': 'total_pages',
        'Total_products': 'total_products',
        'Current_page_scraped': 'current_page_scraped',
        'current_product_scraped': 'current_product_scraped'
    }

    def __init__(self, db: ParseHubDatabase = None):
        """Initialize the service with database connection"""
        self.db = db or ParseHubDatabase()
        self.validation_errors = []
        self.import_stats = {
            'total_records': 0,
            'imported': 0,
            'skipped': 0,
            'errors': []
        }

    def parse_excel_file(self, file_path: str) -> list:
        """
        Parse Excel file and return list of row dictionaries
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List of parsed rows, or empty list on error
        """
        self.validation_errors = []
        
        if not os.path.exists(file_path):
            self.validation_errors.append(f"File not found: {file_path}")
            return []
        
        try:
            # Try pandas first (better for complex operations)
            if pd is not None:
                df = pd.read_excel(file_path, sheet_name=0)
                df = df.fillna('')  # Replace NaN with empty string
                return df.to_dict('records')
            
            # Fallback to openpyxl
            elif openpyxl is not None:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                
                # Get header row
                headers = []
                for cell in ws[1]:
                    headers.append(cell.value)
                
                # Parse data rows
                rows = []
                for row_idx in range(2, ws.max_row + 1):
                    row_dict = {}
                    for col_idx, header in enumerate(headers, 1):
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        row_dict[header] = cell_value or ''
                    rows.append(row_dict)
                
                return rows
            
            else:
                self.validation_errors.append(
                    "Neither pandas nor openpyxl installed. "
                    "Install with: pip install openpyxl pandas"
                )
                return []
                
        except Exception as e:
            self.validation_errors.append(f"Error parsing Excel file: {str(e)}")
            return []

    def validate_metadata_row(self, row: dict) -> tuple[bool, str]:
        """
        Validate a single metadata row
        
        Args:
            row: Dictionary containing row data
            
        Returns:
            Tuple of (is_valid: bool, error_message: str)
        """
        errors = []
        
        # Check required fields
        if not row.get('Personal Project ID') or str(row.get('Personal Project ID')).strip() == '':
            errors.append("Personal Project ID is required")
        
        if not row.get('Project_name') or str(row.get('Project_name')).strip() == '':
            errors.append("Project_name is required")
        
        # Validate numeric fields
        total_pages = row.get('Total_pages')
        if total_pages is not None and total_pages != '':
            try:
                int(total_pages)
            except (ValueError, TypeError):
                errors.append(f"Total_pages must be numeric, got: {total_pages}")
        
        total_products = row.get('Total_products')
        if total_products is not None and total_products != '':
            try:
                int(total_products)
            except (ValueError, TypeError):
                errors.append(f"Total_products must be numeric, got: {total_products}")
        
        current_page = row.get('Current_page_scraped')
        if current_page is not None and current_page != '':
            try:
                int(current_page)
            except (ValueError, TypeError):
                errors.append(f"Current_page_scraped must be numeric, got: {current_page}")
        
        if errors:
            return False, "; ".join(errors)
        
        return True, ""

    def bulk_import_metadata(self, file_path: str, uploaded_by: str = None) -> dict:
        """
        Import metadata records from Excel file into database
        
        Args:
            file_path: Path to Excel file
            uploaded_by: Username of person uploading the file
            
        Returns:
            Dictionary with import statistics and results
        """
        self.import_stats = {
            'total_records': 0,
            'imported': 0,
            'skipped': 0,
            'errors': []
        }
        
        # Create import batch
        file_name = os.path.basename(file_path)
        batch_id = self.db.create_import_batch(file_name, uploaded_by=uploaded_by)
        
        if batch_id is None:
            return {
                'success': False,
                'error': 'Failed to create import batch',
                'stats': self.import_stats
            }
        
        # Parse file
        rows = self.parse_excel_file(file_path)
        
        if not rows:
            return {
                'success': False,
                'error': f'No valid rows found. Errors: {"; ".join(self.validation_errors)}',
                'stats': self.import_stats
            }
        
        self.import_stats['total_records'] = len(rows)
        
        # Import each row
        for row_idx, row in enumerate(rows, start=2):  # Start from 2 because row 1 is header
            # Validate row
            is_valid, error_msg = self.validate_metadata_row(row)
            
            if not is_valid:
                self.import_stats['skipped'] += 1
                self.import_stats['errors'].append({
                    'row': row_idx,
                    'error': error_msg,
                    'personal_id': row.get('Personal Project ID', 'unknown')
                })
                continue
            
            # Extract and clean data
            personal_id = str(row.get('Personal Project ID', '')).strip()
            project_token = str(row.get('Project ID (ParseHub)', '')).strip() or None
            project_name = str(row.get('Project_name', '')).strip()
            region = str(row.get('Region', '')).strip() or None
            country = str(row.get('Country', '')).strip() or None
            brand = str(row.get('Brand', '')).strip() or None
            website_url = str(row.get('Website_url', '')).strip() or None
            last_run_date = self._parse_date(row.get('Last_run_data'))
            
            # Parse numeric fields
            try:
                total_pages = int(row.get('Total_pages')) if row.get('Total_pages') else None
            except (ValueError, TypeError):
                total_pages = None
            
            try:
                total_products = int(row.get('Total_products')) if row.get('Total_products') else None
            except (ValueError, TypeError):
                total_products = None
            
            try:
                current_page = int(row.get('Current_page_scraped', 0)) if row.get('Current_page_scraped') else 0
            except (ValueError, TypeError):
                current_page = 0
            
            try:
                current_product = int(row.get('current_product_scraped', 0)) if row.get('current_product_scraped') else 0
            except (ValueError, TypeError):
                current_product = 0
            
            # Try to link to existing project by token
            project_id = None
            if project_token:
                # Check if project exists in DB
                conn = self.db.connect()
                cursor = conn.cursor()
                cursor.execute('SELECT id FROM projects WHERE token = ?', (project_token,))
                result = cursor.fetchone()
                if result:
                    project_id = result['id']
                self.db.disconnect()
            
            # Add metadata record
            metadata_id = self.db.add_metadata_record(
                personal_project_id=personal_id,
                project_token=project_token,
                project_id=project_id,
                project_name=project_name,
                region=region,
                country=country,
                brand=brand,
                website_url=website_url,
                total_pages=total_pages,
                total_products=total_products,
                import_batch_id=batch_id
            )
            
            if metadata_id:
                # Update progress if current_page is greater than 0
                if current_page > 0 or current_product > 0:
                    self.db.update_metadata_progress(
                        metadata_id,
                        current_page_scraped=current_page,
                        current_product_scraped=current_product,
                        last_run_date=last_run_date
                    )
                
                self.import_stats['imported'] += 1
            else:
                self.import_stats['skipped'] += 1
                self.import_stats['errors'].append({
                    'row': row_idx,
                    'error': 'Failed to insert into database',
                    'personal_id': personal_id
                })
        
        return {
            'success': True,
            'batch_id': batch_id,
            'file_name': file_name,
            'stats': self.import_stats,
            'timestamp': datetime.now().isoformat()
        }

    def _parse_date(self, date_value) -> str:
        """Parse date value to ISO format string"""
        if not date_value:
            return None
        
        if isinstance(date_value, str):
            # Try common date formats
            for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']:
                try:
                    dt = datetime.strptime(date_value, fmt)
                    return dt.isoformat()
                except ValueError:
                    continue
            
            # If no format matched, return as is
            return str(date_value)
        
        # If it's a datetime object
        if hasattr(date_value, 'isoformat'):
            return date_value.isoformat()
        
        return str(date_value) if date_value else None

    def get_import_template(self) -> str:
        """Generate CSV template with expected columns"""
        headers = list(self.EXPECTED_COLUMNS.keys())
        return ','.join(headers)

    def export_import_history(self, limit: int = 100) -> list:
        """Get import batch history"""
        try:
            conn = self.db.connect()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT * FROM import_batches 
                ORDER BY upload_date DESC 
                LIMIT ?
            ''', (limit,))
            
            batches = [dict(row) for row in cursor.fetchall()]
            self.db.disconnect()
            
            return batches
            
        except Exception as e:
            print(f"Error retrieving import history: {e}")
            self.db.disconnect()
            return []


if __name__ == '__main__':
    # Test the service
    service = ExcelImportService()
    
    # Example: Import from sample file
    sample_file = 'd:\\Parsehub\\sample_metadata.xlsx'
    if os.path.exists(sample_file):
        result = service.bulk_import_metadata(sample_file, uploaded_by='test_user')
        print("Import Result:")
        print(json.dumps(result, indent=2))
    else:
        print(f"Sample file not found: {sample_file}")
        print("Template columns:", service.get_import_template())
