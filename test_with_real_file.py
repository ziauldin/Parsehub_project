#!/usr/bin/env python3
import requests
import time
import tempfile
import os

time.sleep(1)

# Create a minimal valid Excel file for testing
# Using a temporary file
print("Testing metadata import with actual Excel file...")

# Create a simple Excel file using openpyxl
try:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Region'
    ws['B1'] = 'Country'
    ws['C1'] = 'Brand'
    ws['A2'] = 'North America'
    ws['B2'] = 'USA'
    ws['C2'] = 'TestBrand'
    
    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
        wb.save(tmp.name)
        temp_path = tmp.name
    
    # Test through frontend proxy
    with open(temp_path, 'rb') as f:
        response = requests.post(
            'http://localhost:3001/api/metadata/import',
            headers={'Authorization': 'Bearer t_hmXetfMCq3'},
            files={'file': ('metadata.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
            timeout=10
        )
    
    print(f'Status: {response.status_code}')
    print(f'Content-Type: {response.headers.get("content-type")}')
    print(f'Response: {response.text[:500]}')
    
    os.unlink(temp_path)
    
except Exception as e:
    print(f'Error: {e}')
